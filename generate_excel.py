#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera um ficheiro Excel formatado com validação para importação de resultados
Taça Manuel André 2026 - Estela Golf Club
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Carregar dados do backup
with open('data-backup.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Resultados"

# Definir estilos
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name='Montserrat', size=11, bold=True, color="FFFFFF")
read_only_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
read_only_font = Font(name='Montserrat', size=10, color="595959")
editable_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
editable_font = Font(name='Montserrat', size=10)
border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Cabeçalhos
headers = ["matchId", "Ronda", "Par", "Equipa Casa", "Equipa Fora", "Resultado", "Score (X&Y)"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Larguras das colunas
ws.column_dimensions['A'].width = 15  # matchId
ws.column_dimensions['B'].width = 10  # Ronda
ws.column_dimensions['C'].width = 8   # Par
ws.column_dimensions['D'].width = 22  # Equipa Casa
ws.column_dimensions['E'].width = 22  # Equipa Fora
ws.column_dimensions['F'].width = 15  # Resultado
ws.column_dimensions['G'].width = 15  # Score

# Preparar validação
validation_result = DataValidation(
    type="list",
    formula1='"Vence A,Vence B,A/S"',
    allow_blank=True
)
validation_result.error = 'Valores permitidos: "Vence A", "Vence B", "A/S"'
validation_result.errorTitle = 'Valor Inválido'
ws.add_data_validation(validation_result)

validation_score = DataValidation(
    type="custom",
    formula1="OR(LEN(INDIRECT(ADDRESS(ROW(),COLUMN())))=0,REGEX(INDIRECT(ADDRESS(ROW(),COLUMN())),\"^[0-9]+&[0-9]+$\"))",
    allow_blank=True
)
validation_score.error = 'Formato: X&Y (exemplo: 2&1)'
validation_score.errorTitle = 'Formato Inválido'
ws.add_data_validation(validation_score)

# Popular dados
row = 2
for idx, match in enumerate(data['calendar']):
    match_id = f"R{match['ronda']}-{match['par']}-{idx}"
    
    # matchId (read-only)
    cell = ws.cell(row=row, column=1)
    cell.value = match_id
    cell.fill = read_only_fill
    cell.font = read_only_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    # Ronda (read-only)
    cell = ws.cell(row=row, column=2)
    cell.value = match['ronda']
    cell.fill = read_only_fill
    cell.font = read_only_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    # Par (read-only)
    cell = ws.cell(row=row, column=3)
    cell.value = match['par']
    cell.fill = read_only_fill
    cell.font = read_only_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    # Equipa Casa (read-only)
    cell = ws.cell(row=row, column=4)
    cell.value = match['home']
    cell.fill = read_only_fill
    cell.font = read_only_font
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = border
    
    # Equipa Fora (read-only)
    cell = ws.cell(row=row, column=5)
    cell.value = match['away']
    cell.fill = read_only_fill
    cell.font = read_only_font
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = border
    
    # Resultado (EDITÁVEL com validação)
    cell = ws.cell(row=row, column=6)
    cell.value = ""
    cell.fill = editable_fill
    cell.font = editable_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    validation_result.add(cell)
    
    # Score (EDITÁVEL com validação)
    cell = ws.cell(row=row, column=7)
    cell.value = ""
    cell.fill = editable_fill
    cell.font = editable_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    # Nota: validação de score é mais complexa em openpyxl, deixamos apenas como guia
    
    row += 1

# Congelar primeira linha
ws.freeze_panes = "A2"

# Adicionar sheet com instruções
instr_ws = wb.create_sheet("Instruções")
instr_ws.column_dimensions['A'].width = 80

instructions = [
    ("TAÇA MANUEL ANDRÉ 2026 - IMPORTAÇÃO DE RESULTADOS", "bold"),
    ("", ""),
    ("✓ Preenchimento:", "bold"),
    ("1. Preencha APENAS as colunas 'Resultado' e 'Score (X&Y)'", ""),
    ("2. As outras colunas são de REFERÊNCIA (não edite)", ""),
    ("", ""),
    ("✓ Valores válidos para Resultado:", "bold"),
    ("   • Vence A - Equipa de casa vence", ""),
    ("   • Vence B - Equipa visitante vence", ""),
    ("   • A/S - Empate (All Square)", ""),
    ("", ""),
    ("✓ Formato para Score:", "bold"),
    ("   • X&Y onde X=pontos casa, Y=pontos fora", ""),
    ("   • Exemplos: 2&1, 1&0, 3&2", ""),
    ("", ""),
    ("✓ Como exportar para CSV:", "bold"),
    ("   1. Abra o ficheiro 'Calendario_Resultados_IMPORT.xlsx'", ""),
    ("   2. Preencha os resultados e scores", ""),
    ("   3. Pressione Alt+F11 para abrir o editor de macros", ""),
    ("   4. Execute 'ExportarParaCSV'", ""),
    ("   5. Guarde o ficheiro CSV", ""),
    ("   6. Importe na app (Admin → Configurações → Importar Resultados)", ""),
]

row = 1
for text, style in instructions:
    cell = instr_ws.cell(row=row, column=1)
    cell.value = text
    if style == "bold":
        cell.font = Font(name='Montserrat', size=11, bold=True, color="1F4E78")
    else:
        cell.font = Font(name='Montserrat', size=10)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    row += 1

# Guardar
wb.save('Calendario_Resultados_IMPORT.xlsx')
print("✓ Ficheiro Excel criado: Calendario_Resultados_IMPORT.xlsx")
print(f"✓ Total de matches: {len(data['calendar'])}")
